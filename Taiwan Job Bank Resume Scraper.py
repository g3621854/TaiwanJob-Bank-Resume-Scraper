import urllib.request
import urllib.parse
import json
import requests
import openpyxl
from lxml import html
import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime
from datetime import datetime, timedelta
import pylightxl as plx
import pandas as pd

"""518_人力爬取"""
def creat_request_vacancies_518(page):
    """ 請求對象的訂製_職缺ID_POST """
    url = 'https://www.518.com.tw/ajax/'
    data = {
        'module': 'comp',
        'action': 'getActiveListInfo',
        'currentLength': page,
        }
    # post請求需編碼及encode
    data = urllib.parse.urlencode(data).encode('utf-8')

    headers = {
        'Cookie':'',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76',
        'X-Requested-With':'XMLHttpRequest',
        }
    request = urllib.request.Request(url=url,headers=headers,data = data)
    return request

def creat_request_resume_518(base_resume):
        """求對象的訂製_所有的應徵名單_GET"""
        headers = {
        'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Connection':'keep-alive',
        'Cookie':'',
        'Host':'www.518.com.tw',
        'Referer':'https://www.518.com.tw/comp-message.html',
        'Sec-Ch-Ua':'"Chromium";v="118", "Microsoft Edge";v="118", "Not=A?Brand";v="99"',
        'Sec-Ch-Ua-Mobile':'?0',
        'Sec-Ch-Ua-Platform':'"Windows"',
        'Sec-Fetch-Dest':'document',
        'Sec-Fetch-Mode':'navigate',
        'Sec-Fetch-Site':'same-origin',
        'Sec-Fetch-User':'?1',
        'Upgrade-Insecure-Requests':'1',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76'
        }
        request = urllib.request.Request(url=base_resume,headers=headers)
        return request

"""1111_人力爬取"""
def creat_request_vacancies_1111():
    """ 請求對象的訂製_職缺ID_GET """
    url = 'https://recruit.1111.com.tw/includes/ajax/getEmpListForPool.ashx?resumeType=1&opened=1'

    headers = {
        'Cookie':'',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76',
        'X-Requested-With':'XMLHttpRequest',
        }
    request = urllib.request.Request(url=url,headers=headers)
    return request

def creat_request_resume_1111(base_resume):
        """求對象的訂製_所有的應徵名單_GET"""
        headers = {
        'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language':'zh-TW,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,ja;q=0.5,zh-CN;q=0.4',
        'Cache-Control':'max-age=0',
        'Connection':'keep-alive',
        'Cookie':'',
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.76'
        }
        request = urllib.request.Request(url=base_resume,headers=headers)
        return request

"""小雞上工_人力爬取"""
def download_resume_chicken():
    # 目標URL和Payload
    url = "https://www.chickpt.com.tw/manager/resume/export"
    payload = {"page": 1, "per_page": 20, "start_date": "", "end_date": "", "jobs": ""}

    # 設置POST請求的標頭
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,ja;q=0.5,zh-CN;q=0.4",
        "Connection": "keep-alive",
        "Content-Length": "64",
        "Content-Type": "application/json;charset=UTF-8",
        # 添加其他必要的標頭，例如Cookie和X-Xsrf-Token
        "Cookie": "",
        "X-Xsrf-Token": "eyJpdiI6IkR5NUNLSHE3b1NqYkZVMnFJdHhGMFE9PSIsInZhbHVlIjoiaHBcL0M0dHlIbkRWNnpqQVNvXC91dlBVd2h6XC9CcW03SW1BcjR3Sk1MaDZVSXhkSWVCWGRWYXdEcndCSnBoZnYyeUdUdklQMkRoUk1aOVRyNVloWERBUmc9PSIsIm1hYyI6IjVlNTQ1ODZiZWU5M2MxMDZkZGE2ZWJjNzgxYzEwNTBiOGMzOTFhMTFiN2U2NTA4MTlmNGJhNjQ3NDg2ZjkxMjUifQ=="
        # 更多標頭...
    }

    # 發送POST請求
    response = requests.post(url, headers=headers, json=payload)

    # 檢查是否成功
    if response.status_code == 200:
        # 將文件內容寫入本地文件
        with open("downloaded_excel.xlsx", "wb") as file:
            file.write(response.content)
        return print("小雞上工文件下載成功。")
    else:
        return print(f"小雞上工文件下載失敗，狀態碼：{response.status_code}")
    

def get_content(request):
    """
    模擬瀏覽器並根據請求獲取內容。
    :param request: 請求對象
    :return: 網頁內容
    """
    try:
        response = urllib.request.urlopen(request)
        return response.read().decode('UTF-8')
    except urllib.error.URLError as e:
        print(f"URL 錯誤: {e}")
        return None


def parse_application_dates_518(application_date):
    """將應徵日期轉換為標準日期格式"""
    converted_dates = []
    for date in application_date:
        if date == "今日":
            formatted_date = datetime.now().strftime("%Y/%m/%d")
        elif date == "昨日":
            formatted_date = (datetime.now() - timedelta(days=1)).strftime("%Y/%m/%d")
        elif "天前" in date:
            days = int(re.search(r'(\d+)天前', date).group(1))
            formatted_date = (datetime.now() - timedelta(days=days)).strftime("%Y/%m/%d")
        else:
            # 只有當日期不是上述情況時才進行替換操作
            formatted_date = date.replace('-', '/')

        converted_dates.append(formatted_date)
    return converted_dates

def extract_applicant_info_518(tree):
    """提取應徵者資訊"""
    job_title = tree.xpath('//li[@class="tag"]/text()')
    classification, store, store_class = extract_job_tittle_tree(tree,job_title)
    name = [element.text.strip() for element in tree.xpath('//div[@class="name-inbox"]/a')]
    sex = tree.xpath('//div[@class="list-center-box"]//div[@class="info-inbox"]/p[1]/text()')
    age = [int(element.replace('歲', '')) for element in tree.xpath('//div[@class="list-center-box"]//div[@class="info-inbox"]/p[2]/text()')]
    # 提取應徵日期並轉換
    application_date = tree.xpath('//div[@class="list-center-box"]//p[@class="date"]/span/text()')
    converted_dates = parse_application_dates_518(application_date)
    residence = tree.xpath('//li[@class="area"]/p/text()')
    education_levels = extract_education_levels_518(tree)
    education_school = [text.split(' / ')[0] for text in tree.xpath('//div[@class="list-center-box"]//li[@class="school"]/p[1]/text()')]
    education_department = [text.split(' / ')[-1] for text in tree.xpath('//div[@class="list-center-box"]//li[@class="school"]/p[1]/text()')]
    work_experience = tree.xpath('//div[@class="info-inbox"]/p[4]/text()')#尚未處理
    last_job_list = extract_last_job_518(tree)
    job_status = tree.xpath('//div[@class="list-center-box"]//div[@class="info-inbox"]/p[5]/text()')

    return job_title, classification, store, store_class, name, sex, age, converted_dates, residence, education_levels, education_school, education_department, work_experience, last_job_list, job_status

def extract_education_levels_518(tree):
    """抓取最高學歷"""
    info_boxes = tree.xpath('//div[@class="info-inbox"]')
    education_levels=[]
    for box in info_boxes:
    # 嘗試抓取每個盒子的第三個段落
        education = box.xpath('.//p[3]/text()')
        # 檢查是否抓到資料
        if education:
            # 如果有資料，加入到列表
            education_levels.append(education[0])
        else:
            # 如果沒有資料（即空的 p 標籤），加入空字串
            education_levels.append('')
    return education_levels

def extract_last_job_518(tree):
    """提取應徵者的最後一份工作"""
    last_job_list = []
    lxml_job = tree.xpath('//li[@class="job"]')
    for element in lxml_job:
        last_job = element.xpath('./p[1]/text()')
        if last_job:
            job_text = last_job[0].split('/')[1].strip() if '/' in last_job[0] else last_job[0]
            last_job_list.append(job_text)
        else:
            last_job_list.append('')
    return last_job_list

def extract_job_tittle_tree(tree,job_title):
    """提取分店及班別"""
   # 预编译正则表达式
    pattern_store = re.compile(r'門市(.*?)-')
    pattern_brackets = re.compile(r'\((.*?)\)')
    pattern_after_dash = re.compile(r'-(.*?)$')
    pattern_after_parenthesis = re.compile(r'\)(.*?)$')

    store = []
    store_class = []
    classification = []

    for title in job_title:
        if '【' in title:
            match = pattern_store.search(title)
            store.append(match.group(1) if match else "")

            match = pattern_after_dash.search(title)
            store_class.append(match.group(1).strip() if match else "")

            classification.append("中央幹部")
        elif "-儲備幹部" in title:
            store.append("儲備幹部")

            match = pattern_brackets.search(title)
            store_class.append(match.group(1) if match else "")

            classification.append("地方儲備")
        else:
            match = pattern_brackets.search(title)
            store.append(match.group(1) if match else "")

            match = pattern_after_parenthesis.search(title)
            clean_text = re.sub(r'店員', '', match.group(1)) if match else ""
            store_class.append(clean_text)

            classification.append("一般店員")
    return classification, store, store_class


def resume_df_518(content_resume):
    tree = html.fromstring(content_resume) # 使用lxml解析網頁內容
    # 提取應徵者資訊
    job_title, classification, store, store_class, name, sex, age, converted_dates, residence, education_levels, education_school, education_department, work_experience, last_job_list, job_status = extract_applicant_info_518(tree)
    track = ['T-尚未聯繫']
    joining_date = ['']
    memo = ['']
    new_data = {
            # '職缺名稱': job_title * len(name),
            '職缺分類':classification*len(name),
            '分店': store* len(name),
            '班別':store_class* len(name),
            '姓名': name,
            '性別': sex,
            '年紀': age,
            '追蹤狀態':track* len(name),
            '應徵日期': converted_dates,
            '入職日期':joining_date* len(name),
            '備註': memo* len(name),
            '居住地' : residence,
            '最高學歷': education_levels,
            '學校名稱': education_school,
            '科系名稱': education_department,
            '工作經歷': work_experience,
            '上份工作': last_job_list,
            '職業狀態': job_status,
            }

    # 合并new_data到applicant_data
    for key, value in new_data.items():
        if key in applicant_data_518:
            applicant_data_518[key].extend(value)
        else:
            applicant_data_518[key] = value
    # print(applicant_data_518)

    return applicant_data_518


def find_ID_518(content):
    """拿到職缺url"""
    id_data = {}

    # 轉換為列表
    id_list = json.loads(content)
   # 遍歷列表並將每個字典添加到一個新的字典中
    for item in id_list:
        id_data[item['id']] = item
    # 遍歷所有職缺ID job_data:str
    for job_data in id_data:
        base_resume = 'https://www.518.com.tw/comp-active.html?id='+job_data
        print(f'URL:{base_resume}')
         # 請求對象的訂製_履歷
        request_resume = creat_request_resume_518(base_resume)
        # 模擬瀏覽器並獲取履歷數據
        content_resume = get_content(request_resume)
        # 抓取data數據
        applicant_data = resume_df_518(content_resume)
        # print(applicant_data)
    return applicant_data


def parse_518_page():
    """爬取518主程式"""
    df_518 = pd.DataFrame()
    for page in [0,50]:
        if page == 0 or page == 50:
            # 請求對象的訂製_職缺
            request = creat_request_vacancies_518(page)
            # 模擬瀏覽器並獲取職缺id數據
            content_va = get_content(request)
            # 拿到職缺url及履歷清單
            applicant_data = find_ID_518(content_va)
            # print(applicant_data)
            df_518 = pd.DataFrame(applicant_data)
    print(applicant_data)
    df_518['平台'] = '518'
    return df_518

"""yes123_人力爬取"""

def create_request_123(page):
    """
    創建網絡請求對象。(post)
    :param page: 頁碼
    :return: 請求對象
    """
    base_url = 'https://ent.yes123.com.tw/admin/corp/resume/resume_vip.asp?pn=13&t_type=33'
    data = {
        'pageNO': page,
        'page_size': 20,
    }
    data_encoded = urllib.parse.urlencode(data).encode('UTF-8')
    headers = {
        'Cookie': '_gcl_au=1.1.1465618373.1699627043; _fbp=fb.2.1699627043725.1988932145; _hjFirstSeen=1; _hjIncludedInSessionSample_3426572=0; _hjSession_3426572=eyJpZCI6IjNlMmMzZWYyLTliNjgtNDQxYi04ZmFhLTY5YzVhZGNkOTg5ZSIsImNyZWF0ZWQiOjE2OTk2MjcwNDM4MDUsImluU2FtcGxlIjpmYWxzZSwic2Vzc2lvbml6ZXJCZXRhRW5hYmxlZCI6dHJ1ZX0=; _hjAbsoluteSessionInProgress=0; isfindex=1; ASPSESSIONIDCWTSSRSB=EBJODBGCANBLACPENPCCHHEG; ASP.NET_SessionId=638838783; StepCookie_id=638838783; ClientIP=111.254.36.168; __lt__cid=6def48f1-eb2c-4049-a9f6-3142e0a31eb4; _hjSessionUser_3426572=eyJpZCI6ImIzODM2OThkLWY4NzEtNWRmYy04NTY4LTg5ODRkN2U3OGQ1YiIsImNyZWF0ZWQiOjE2OTk2MjcwNDM4MDMsImV4aXN0aW5nIjp0cnVlfQ==; dcard-adkt-device=07b674de-933c-4ebd-b2ac-eaf5b4bf8240; _gid=GA1.3.211979639.1699627059; _ga_B6PF5H3656=GS1.3.1699627059.1.0.1699627059.60.0.0; _ga=GA1.1.1920163115.1699627043; yes123_make_cookie=ab71b7796b7ded930f27bc9c414f92fa; ck%5Fp%5Fkey=36%3D%3D6%3D2%3D6; ck%5Fp%5Fid=75754450460455%5F61%3C26234; ck%5Fp%5Fsub%5Fid=75754450460455%5F61%3C26234%5F7%3D%3C4051; c%5Fp%5Fstaffing=%5F%5F%5F0%5F%5F%5F0%5F%5F%5F0; today%5Fbonus=; sift%5Fwy%5Fyear=; p%5Fservice%5Fsa=%E9%BB%83%E6%80%A1%E7%8F%8A%2A%2A%2A02%2D26560123%232725%2A%2A%2Asunny%5Fhuang%40yes123%2Ecom%2Etw%2A%2A%2A; sift%5Fyear=; pd=nguvna; sift%5Fsubja=; sift%5Ftxkey=; p%5Fsa=%E8%94%A1%E7%BE%8E%E6%83%A0%2A%2A%2A02%2D26560123%232798%2A%2A%2Ajanice%5Ftsai%40yes123%2Ecom%2Etw%2A%2A%2A; cpid=20201105135100%5F34973761; sift%5Fzw=; c%5Fpdata=%E8%A9%B9%E5%A8%89%E5%A6%A4%5F%5F%5F%5F%5F%5F; pu=%7Fpu46414%3C; apikey=; people%5Fuser%5Fid=; pi=20201105135100%5F34973761; cpsubid=20201105135100%5F34973761%5F2891504; apikeyweb=5883557028c062eca2af5a739288e5df; pua=yes; sift%5Fpeople%5Fnum=; sift%5Fj%5Fs=; today%5Fthis%5Fbonus=0; LLT=2023%2D11%2D10+22%3A38%3A08%2E933; _ga_14SNTQ4DKL=GS1.1.1699627043.1.1.1699627811.14.0.0; step=20; citrix_ns_id=vcXbvA07NSoEqyP5ylR6FGlQ4hA0000',
        'Referer': 'https://ent.yes123.com.tw/admin/corp/resume/resume_vip.asp?pn=13',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0'
   }

    return urllib.request.Request(url=base_url, headers=headers, data=data_encoded)



def extract_job_status_123(tree):
    """提取職業狀態

    Args:
        tree (str): 網頁內容

    Returns:
        list: 工作狀態清單
    """
    last_job_list = ['']
    job_status_list = []
    info_boxes = tree.xpath('//tr[not(.//td[contains(text(), "關閉履歷")])]//td[@class="pmg show_resume"]')
    for box in info_boxes:
    # 嘗試抓取每個盒子的第三個段落
       job_status = box.xpath('./div/div[1]/text()')
       # 檢查是否抓到資料
       if job_status:
           # 如果有資料，加入到列表
           job_status_list.append(job_status[0])
       else:
           # 如果沒有資料（即空的 p 標籤），加入空字串
           job_status_list.append('')
    return job_status_list

def parse_content_123(content):
    """
    解析網頁內容。
    :param content: 網頁內容
    :return: 解析後的數據
    """
    tree = html.fromstring(content)
    job_title = tree.xpath('//tr[not(.//td[contains(text(), "關閉履歷")])]//td[@class="show_resume"]//span/text()')
    classification, store, store_class = extract_job_tittle_tree(tree,job_title)
    name = tree.xpath('//tr[not(.//td[contains(text(), "關閉履歷")])]//div[@class="left t_left s_ml15"]/a/text()')
    sex = [text.split('／')[-1] for text in tree.xpath('//tr[not(.//td[contains(text(), "關閉履歷")])]//div[@class="left t_left s_ml15"]/text()')]
    age = [int(text.split('／')[0].replace('歲', '') )for text in tree.xpath('//tr[not(.//td[contains(text(), "關閉履歷")])]//div[@class="left t_left s_ml15"]/text()')]
    track = ['T-尚未聯繫']
    converted_dates = [element.replace('.','/') for element in tree.xpath('//tr[not(.//td[contains(text(), "關閉履歷")])]//font/p/text()')]
    joining_date = ['']
    memo = ['']
    residence = ['']
    education_levels = ['']
    education_school = tree.xpath('//tr/td[@width="18%" and @class="show_resume"]/text()[1]')
    education_department =  tree.xpath('//tr/td[@width="18%" and @class="show_resume"]/text()[2]')
    work_experience = ['']
    last_job_list = ['']
    job_status = extract_job_status_123(tree)

    new_data = {
            # '職缺名稱': job_title * len(name),
            '職缺分類':classification,
            '分店': store,
            '班別':store_class,
            '姓名': name,
            '性別': sex,
            '年紀': age,
            '追蹤狀態':track* len(name),
            '應徵日期': converted_dates,
            '入職日期':joining_date* len(name),
            '備註': memo* len(name),
            '居住地' : residence* len(name),
            '最高學歷': education_levels* len(name),
            '學校名稱': education_school,
            '科系名稱': education_department,
            '工作經歷': work_experience* len(name),
            '上份工作': last_job_list* len(name),
            '職業狀態': job_status,
            }
        # 合并new_data到applicant_data
    for key, value in new_data.items():
        if key in applicant_data_123:
            applicant_data_123[key].extend(value)
        else:
            applicant_data_123[key] = value

    return applicant_data_123


def parse_yes123_page():
    applicant_data = {}
    for page in range(1, 15):
        request = create_request_123(page)
        print(request)
        content = get_content(request)
        applicant_data = parse_content_123(content)
        df_123 = pd.DataFrame(applicant_data)
    print(applicant_data)
    df_123 = df_123.drop_duplicates(subset='姓名', keep=False) # 去重複
    df_123['平台'] = 'yes123'
    return df_123

def extract_applicant_info_1111(tree,job_title):
    """提取履歷訊息"""
    classification, store, store_class = extract_job_tittle(job_title)
    name = [element.text for element in tree.xpath('//li[@class="name noborder"]//font')]
    sex = extract_sex(tree)
    age_elements = tree.xpath('//li[@class="name noborder"]/text()')
    # 使用正則表達式提取數字，僅當匹配存在時添加到列表
    age = [int(match.group()) for age in age_elements if (match := re.search(r'\d+', age))]
    converted_dates = tree.xpath('//li[@class="view"]/@title')
    education_levels = [element.text for element in tree.xpath('//ul[contains(@class, "Areabox_ListB")]//li[@class="edu"]')]
    education_school = [element.text for element in tree.xpath('//ul[contains(@class, "Areabox_ListB")]//li[@class="school"]')]
    education_department = [element.text for element in tree.xpath('//ul[contains(@class, "Areabox_ListB")]//li[@class="major"]')]
    work_experience, last_job_list = extract_work_experience_elements(tree)
    job_status = [element.text for element in tree.xpath('//ul[contains(@class, "Areabox_ListB")]//li[@class="ing"]')]
    return  classification, store, store_class, name, sex, age, converted_dates,  education_levels, education_school, education_department, work_experience, last_job_list, job_status

def extract_job_tittle(job_title):
    """提取分店及班別"""
   # 预编译正则表达式
    pattern_store = re.compile(r'門市(.*?)-')
    pattern_brackets = re.compile(r'\((.*?)\)')
    pattern_after_dash = re.compile(r'-(.*?)$')
    pattern_after_parenthesis = re.compile(r'\)(.*?)$')

    store = []
    store_class = []
    classification = []
    title = job_title

    if '【' in title:
            match = pattern_store.search(title)
            store.append(match.group(1) if match else "")

            match = pattern_after_dash.search(title)
            store_class.append(match.group(1).strip() if match else "")

            classification.append("中央幹部")
    elif "-儲備幹部" in title:
            store.append("儲備幹部")

            match = pattern_brackets.search(title)
            store_class.append(match.group(1) if match else "")

            classification.append("地方儲備")
    else:
            match = pattern_brackets.search(title)
            store.append(match.group(1) if match else "")

            match = pattern_after_parenthesis.search(title)
            clean_text = re.sub(r'店員', '', match.group(1)) if match else ""
            store_class.append(clean_text)

            classification.append("一般店員")
    return classification, store, store_class

def extract_sex(tree):
    """性別"""
    sex_elements = tree.xpath('//li[@class="name noborder"]/em[contains(@class, "sex")]')
    sex_classes = [element.get('class') for element in sex_elements]
    # # 创建一个空列表来保存性别信息
    sex = []
    # 遍历找到的元素
    for sex_class in sex_classes:
        # 根据类名确定性别
        if 'sex1' in sex_class:
            sex.append('男')
        else:
            sex.append('女')
    return sex

def extract_work_experience_elements(tree):
    """提取工作經歷、上份工作"""
    work_experience_elements = tree.xpath('//*[@class="grow"] | //*[@class="grow "]')
    # 遍历找到的元素并打印包含 "累计年资" 的文本
    work_experience = []
    last_job_list = []

    for w_element in work_experience_elements:
         # 输出文本内容
        work_experience_list = w_element.text
        # 检查是否含有“無工作經驗”
        if '累計年資：無工作經驗' in work_experience_list:
            last_job_list.append('無工作經驗')
            work_experience.append('無工作經驗')
        else:
            # 用於提取“累計年資”
            if '累計年資' in work_experience_list:
                year_exp = work_experience_list.split('累計年資：')[1].split('\n')[0]
                # print(year_exp)
                work_experience.append(work_experience_list.split('累計年資：')[1])
            # 用於提取“前一工作”
            if '前一工作' in work_experience_list:
                match = re.search(r'：.*? / (.*?) \d{4}/\d{2}~', work_experience_list)
                if match:
                    last_job_list.append(match.group(1))
                else:
                    last_job_list.append('未知職位')
    return(work_experience, last_job_list)

def resume_df_1111(content_resume,job_title_hp):
    """
    從當前頁面提取履歷資料。
    :param driver: 瀏覽器驅動實例。
    :return: 返回一個字典，包含從頁面提取的履歷資料。
    """
    tree = html.fromstring(content_resume)
    classification, store, store_class, name, sex, age, converted_dates, education_levels, education_school, education_department, work_experience, last_job_list, job_status = extract_applicant_info_1111(tree,job_title_hp)
    track = ['T-尚未聯繫']
    joining_date = ['']
    memo = ['']
    residence = ['']

    new_data = {
            # '職缺名稱': job_title * len(name),
            '職缺分類':classification* len(name),
            '分店': store* len(name),
            '班別':store_class* len(name),
            '姓名': name,
            '性別': sex,
            '年紀': age,
            '追蹤狀態':track* len(name),
            '應徵日期': converted_dates,
            '入職日期':joining_date* len(name),
            '備註': memo* len(name),
            '居住地' : residence* len(name),
            '最高學歷': education_levels,
            '學校名稱': education_school,
            '科系名稱': education_department,
            '工作經歷': work_experience,
            '上份工作': last_job_list,
            '職業狀態': job_status,
            }
        # 合并new_data到applicant_data
    # print(new_data)
    for key, value in new_data.items():
        if key in applicant_data_1111:
            applicant_data_1111[key].extend(value)
        else:
            applicant_data_1111[key] = value

    return applicant_data_1111



def find_ID_1111(content):
    """拿到職缺url"""
        # 轉換為列表
    id_list = json.loads(content)
    job_titles =  [str(item["Item2"]) for item in id_list]
    id_data = [str(item["Item1"]) for item in id_list]
    # job_title_hp = job_title[12]

    # 同步遍歷兩個列表
    for job_data, job_title in zip(id_data, job_titles):
        base_resume = 'https://recruit.1111.com.tw/ResumePoolActive.aspx?EmpNo='+ job_data +'#opened&eNo='+ job_data
        # print(job_title)
        print(f'base_resume:{base_resume}')
        # 請求對象的訂製_履歷
        request_resume = creat_request_resume_1111(base_resume)
        # 模擬瀏覽器並獲取履歷數據
        content_resume = get_content(request_resume)
        # print(content_resume)
        # # # 抓取data數據
        applicant_data = resume_df_1111(content_resume,job_title)

    return applicant_data


def parse_1111_page():
    """爬取1111主程式"""

    # 請求對象的訂製_職缺
    request = creat_request_vacancies_1111()
    # 模擬瀏覽器並獲取職缺id數據
    content_va = get_content(request)
    # # 拿到職缺url及履歷清單
    applicant_data = find_ID_1111(content_va)

    print(applicant_data)
    df_1111 = pd.DataFrame(applicant_data)

    df_1111['平台'] = '1111'

    return df_1111

def determine_education_level(school):
    if '大學' in school or '學院' in school or '科技大學' in school or '餐旅' in school:
        return '大學'
    elif '國中' in school:
        return '國中'
    else:
        return '高中職'

def data_conversion_ch():
    # 讀取 Excel 檔案並初始化 DataFrame
    file_path = "C:\\Users\\admin\\Desktop\\Work\\003_履歷整理\\downloaded_excel.xlsx"
    db = plx.readxl('downloaded_excel.xlsx')
    name_first_sheet = db.ws_names[0]
    sheet_data = list(db.ws(ws=name_first_sheet).rows)
    excel_data = pd.DataFrame(data=sheet_data[1:], columns=sheet_data[0])

    # Selecting the required columns
    required_columns = ['應徵時間', '應徵職缺', '姓名', '性別', '年齡', '居住縣市', '居住地區', '就讀學校', '就讀科系','工作經驗 - 職務名稱']
    applicant_df = excel_data[required_columns]

    # Cleaning the DataFrame to remove rows where essential fields like name are missing
    applicant_df = applicant_df[applicant_df['姓名'] != ""]
    # display(applicant_df)

    applicant_df['居住地'] = applicant_df['居住縣市'] + applicant_df['居住地區']

    # 整理並創建新的欄位
    applicant_df['最高學歷'] = applicant_df['就讀學校'].apply(determine_education_level)
    applicant_df['應徵時間'] = pd.to_datetime(applicant_df['應徵時間']).dt.date

    # 藉由職缺名稱進行資料分類
    classification, store, store_class = extract_job_tittle(applicant_df['應徵職缺'])
    applicant_df['職缺分類'], applicant_df['分店'], applicant_df['班別'] = classification, store, store_class

    # 刪除不需要的欄位並重新命名
    applicant_df.drop(['應徵職缺', '居住縣市', '居住地區'], axis=1, inplace=True)
    applicant_df.rename(columns={'就讀學校': '學校名稱', '就讀科系': '科系名稱','工作經驗 - 職務名稱':'上份工作','應徵時間':'應徵日期','年齡':'年紀'}, inplace=True)
    applicant_df['應徵日期'] = applicant_df['應徵日期'].astype(str).replace('-', '/', regex=True)

    # 新增預設欄位
    applicant_df['追蹤狀態'] = "T-尚未聯繫"
    applicant_df['平台'] = "小雞上工"
    applicant_df['入職日期'] = ""
    applicant_df['備註'] = ""
    applicant_df['工作經歷'] = ""
    applicant_df['職業狀態'] = ""

    # Reordering the columns as per the specified order
    column_order = [
        '職缺分類', '分店', '班別', '姓名', '性別', '年紀', 
        '追蹤狀態', '應徵日期', '入職日期', '備註', '居住地', 
        '最高學歷', '學校名稱', '科系名稱', '工作經歷', '上份工作', 
        '職業狀態', '平台'
    ]
    applicant_ch = applicant_df[column_order]

    # Displaying the updated DataFrame
    applicant_ch.head()
    excel_file = 'df_chicken.xlsx'  # 替換為你的 Excel 文件路徑
    applicant_ch.to_excel(excel_file, index=False)
    return applicant_ch
def date_conversion(combined_data):
    """將日期格式格式化 避免被excel自動改成其他格式"""
    # combined_data['應徵日期'] = pd.to_datetime(combined_data['應徵日期'], errors='coerce', format='%Y/%m/%d')

    combined_data['應徵日期'] = pd.to_datetime(combined_data['應徵日期'], errors='coerce')
    # 檢查 '應徵日期' 是否為 datetime 對象
    if pd.api.types.is_datetime64_any_dtype(combined_data['應徵日期']):
#       combined_data['應徵日期'] = combined_data['應徵日期'].dt.strftime('%Y/%m/%d')
      combined_data['應徵日期'] = pd.to_datetime(combined_data['應徵日期'], errors='coerce', format='%Y/%m/%d')

    return combined_data

def import_data():
    # 讀取舊資料和新資料
    old_data = pd.read_excel('履歷格式_測試.xlsx',sheet_name='應徵履歷',header=1)
    new_data = pd.read_excel('df_unique.xlsx',sheet_name='Sheet1')

    # 指定用於判斷重複的列名
    # 這裡我們假設使用'姓名'和'應徵日期'兩列來判斷是否重複
    columns_to_check = ['姓名', '應徵日期']

    # 合併資料
    combined_data = pd.concat([old_data, new_data]) # 合併成功
    # 將應徵日期格式化成datetime
#     combined_data = date_conversion(combined_data)
    combined_data['應徵日期'] = pd.to_datetime(combined_data['應徵日期'], errors='coerce', format='%Y/%m/%d')

    # 根據特定列去除重複紀錄
    # keep='first'表示保留第一次出現的紀錄
    unique_data = combined_data.drop_duplicates(subset=columns_to_check, keep='first')
    df_unique_data = pd.DataFrame(unique_data)

    # 重設索引，確保從0開始且沒有間隔
    df_unique_data.reset_index(drop=True, inplace=True)

    # 下載檔案
    excel_file = 'df_unique_data.xlsx'
    df_unique_data.to_excel(excel_file, index=False)

    # 加載現有的Excel檔案
    book = load_workbook('履歷格式_空白.xlsx')
    sheet = book['應徵履歷']  # 或者其他您想要寫入數據的工作表

    # 寫入列標題
    headers_row = 2  # 假設列標題在第三行
    headers = ['職缺分類','分店','班別', '姓名','性別','年紀','追蹤狀態','應徵日期','入職日期','備註','居住地','最高學歷','學校名稱','科系名稱','工作經歷','上份工作','職業狀態','平台']  # 根據您的需求添加或修改
    for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=headers_row, column=col_num, value=header)

            # 寫入數據
    for index, row in df_unique_data.iterrows():
            sheet.cell(row=index+3, column=1, value=row['職缺分類'])
            sheet.cell(row=index+3, column=2, value=row['分店'])
            sheet.cell(row=index+3, column=3, value=row['班別'])
            sheet.cell(row=index+3, column=4, value=row['姓名'])
            sheet.cell(row=index+3, column=5, value=row['性別'])
            sheet.cell(row=index+3, column=6, value=row['年紀'])
            sheet.cell(row=index+3, column=7, value=row['追蹤狀態'])
            sheet.cell(row=index+3, column=8, value=row['應徵日期'])
            sheet.cell(row=index+3, column=9, value=row['入職日期'])
            sheet.cell(row=index+3, column=10, value=row['備註'])
            sheet.cell(row=index+3, column=11, value=row['居住地'])
            sheet.cell(row=index+3, column=12, value=row['最高學歷'])
            sheet.cell(row=index+3, column=13, value=row['學校名稱'])
            sheet.cell(row=index+3, column=14, value=row['科系名稱'])
            sheet.cell(row=index+3, column=15, value=row['工作經歷'])
            sheet.cell(row=index+3, column=16, value=row['上份工作'])
            sheet.cell(row=index+3, column=17, value=row['職業狀態'])
            sheet.cell(row=index+3, column=18, value=row['平台'])
    # 保存工作簿
    book.save('履歷格式_測試.xlsx')

if __name__ == '__main__':
    applicant_data_518 = {}
    applicant_data_123 = {}
    applicant_data_1111 = {}
    download_resume_chicken()
    df_1111 = parse_1111_page()
    df_518 = parse_518_page()
    df_123 = parse_yes123_page()
    df_chicken = data_conversion_ch()
    df_unique = pd.concat([df_518, df_123, df_1111, df_chicken])
    #  下載成excel檔案
    excel_file = 'df_unique.xlsx'  # 替換為你的 Excel 文件路徑
    df_unique.to_excel(excel_file, index=False)
    import_data()